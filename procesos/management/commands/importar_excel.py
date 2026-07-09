"""
Importa el Excel oficial de procesos judiciales (una sola hoja con
secciones apiladas por categoría, un proceso por fila) a la base de
datos normalizada.
"""

import re

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from catalogos.models import Categoria, EstadoProceso, Juzgado, Parte, TipoProceso
from procesos.models import DetalleContrato, HistorialEstado, Proceso, ProcesoParte

ESTADO_POR_DEFECTO = "En trámite"
GADP_NOMBRE = "Gobierno Autónomo Departamental de Potosí"

SECCIONES = [
    ("PROCESOS CONTENCIOSOS", "Contencioso", "contencioso"),
    ("PROCESOS COACTIVOS SOCIALES", "Coactivo social", "generico"),
    ("PROCESOS COACTIVOS", "Coactivo fiscal", "generico"),
    ("PROCESOS CIVILES", "Civil", "generico"),
    ("PROCESOS LABORALES", "Laboral", "generico"),
    ("PROCESOS AGROAMBIENTALES", "Agroambiental", "generico"),
    ("PROCESOS CONSTITUCIONALES", "Constitucional", "generico"),
]


def limpiar_texto(valor):
    if valor is None:
        return ""
    return re.sub(r"\s+", " ", str(valor)).strip(' "\xa0')


def a_nurej(valor):
    texto = limpiar_texto(valor)
    if not texto or texto in {"-", "--"}:
        return None
    return texto[:50]


def detectar_titulo_seccion(fila):
    for celda in fila[:2]:
        if not celda or not isinstance(celda, str):
            continue
        texto = limpiar_texto(celda).upper()
        for titulo, categoria, _ in SECCIONES:
            if texto == titulo:
                return categoria
    return None


def es_fila_total_o_vacia(fila):
    if fila is None or all(c is None for c in fila):
        return True
    for celda in fila:
        if isinstance(celda, str) and limpiar_texto(celda).upper().startswith("TOTAL"):
            return True
    return False


def es_fila_encabezado(fila):
    return len(fila) > 1 and isinstance(fila[1], str) and limpiar_texto(fila[1]).upper() == "Nº"


class Command(BaseCommand):
    help = "Importa el Excel oficial de procesos judiciales a la base de datos."

    def add_arguments(self, parser):
        parser.add_argument("excel_path", type=str)
        parser.add_argument("--hoja", type=str, default=None)
        parser.add_argument("--dry-run", action="store_true")

    def handle(self, *args, **options):
        try:
            wb = openpyxl.load_workbook(options["excel_path"], data_only=True)
        except FileNotFoundError as exc:
            raise CommandError(f"No se encontró el archivo: {options['excel_path']}") from exc

        ws = wb[options["hoja"]] if options["hoja"] else wb[wb.sheetnames[0]]
        filas = list(ws.iter_rows(values_only=True))

        estado_defecto = EstadoProceso.objects.filter(nombre=ESTADO_POR_DEFECTO).first()
        if not estado_defecto:
            raise CommandError("No existe el estado 'En trámite'. Corre primero: python manage.py seed_catalogos")

        # --- Primera pasada: leer todas las filas y agruparlas por (categoría, NUREJ) ---
        categoria_actual = None
        formato_actual = "generico"
        grupos = {}       # (categoria, nurej) -> lista de filas
        sin_nurej = []     # filas sin NUREJ, cada una es su propio proceso

        for fila in filas:
            nombre_categoria = detectar_titulo_seccion(fila)
            if nombre_categoria:
                categoria_actual, _ = Categoria.objects.get_or_create(nombre=nombre_categoria)
                formato_actual = next(f for _, cat, f in SECCIONES if cat == nombre_categoria)
                continue

            if categoria_actual is None:
                continue
            if es_fila_encabezado(fila) or es_fila_total_o_vacia(fila):
                continue

            nro = fila[1] if len(fila) > 1 else None
            if nro in (None, ""):
                continue

            nurej_limpio = a_nurej(fila[6] if len(fila) > 6 else None)
            if nurej_limpio:
                clave = (categoria_actual.id, nurej_limpio)
                grupos.setdefault(clave, {"categoria": categoria_actual, "formato": formato_actual, "filas": []})
                grupos[clave]["filas"].append(fila)
            else:
                sin_nurej.append({"categoria": categoria_actual, "formato": formato_actual, "fila": fila})

        # --- Segunda pasada: crear un Proceso por grupo (fusiona los que compartían NUREJ) ---
        total_creados = 0
        total_saltados = 0
        total_fusionados = 0

        with transaction.atomic():
            sp = transaction.savepoint()

            for clave, grupo in grupos.items():
                try:
                    self._crear_proceso(grupo["filas"], grupo["categoria"], grupo["formato"], estado_defecto)
                    total_creados += 1
                    if len(grupo["filas"]) > 1:
                        total_fusionados += 1
                except Exception as exc:
                    total_saltados += 1
                    nro_ref = grupo["filas"][0][1]
                    self.stderr.write(self.style.WARNING(
                        f"Grupo saltado ({grupo['categoria'].nombre} Nº {nro_ref}, NUREJ {clave[1]}): {exc}"
                    ))

            for item in sin_nurej:
                try:
                    self._crear_proceso([item["fila"]], item["categoria"], item["formato"], estado_defecto)
                    total_creados += 1
                except Exception as exc:
                    total_saltados += 1
                    nro_ref = item["fila"][1]
                    self.stderr.write(self.style.WARNING(
                        f"Fila saltada ({item['categoria'].nombre} Nº {nro_ref}): {exc}"
                    ))

            if options["dry_run"]:
                transaction.savepoint_rollback(sp)
                self.stdout.write(self.style.WARNING("Dry-run: no se guardó nada."))
            else:
                transaction.savepoint_commit(sp)

        self.stdout.write(self.style.SUCCESS(
            f"Procesos creados: {total_creados}  |  Fusionados por NUREJ compartido: {total_fusionados}  |  Saltados: {total_saltados}"
        ))

    def _crear_proceso(self, filas_grupo, categoria, formato, estado_defecto):
        primera = list(filas_grupo[0]) + [None] * (9 - len(filas_grupo[0]))
        _, nro, demandante, proyecto_o_motivo, _demandado, juzgado_txt, nurej, estado_txt, profesional = primera[:9]

        juzgado, _ = Juzgado.objects.get_or_create(nombre=limpiar_texto(juzgado_txt)[:255] or "Sin especificar")
        nurej_limpio = a_nurej(nurej)

        proceso = Proceso.objects.create(
            nro_correlativo=str(nro),
            nurej=nurej_limpio,
            categoria=categoria,
            juzgado=juzgado,
            estado_actual=estado_defecto,
            abogado_referencia=limpiar_texto(profesional)[:255],
        )

        nombre_demandante = limpiar_texto(demandante)
        if nombre_demandante:
            self._vincular_partes(proceso, nombre_demandante, ProcesoParte.Rol.ACTIVA)

        if formato == "contencioso":
            parte, _ = Parte.objects.get_or_create(nombre=GADP_NOMBRE)
            ProcesoParte.objects.create(proceso=proceso, parte=parte, rol=ProcesoParte.Rol.PASIVA)
            if limpiar_texto(proyecto_o_motivo):
                DetalleContrato.objects.create(proceso=proceso, proyecto=limpiar_texto(proyecto_o_motivo)[:500])
        else:
            # Un demandado por cada fila del grupo (así se fusionan los litisconsorcios
            # que en el Excel quedaron repartidos en varias filas con el mismo NUREJ).
            for fila in filas_grupo:
                fila_completa = list(fila) + [None] * (9 - len(fila))
                demandado_fila = limpiar_texto(fila_completa[4])
                if demandado_fila:
                    self._vincular_partes(proceso, demandado_fila, ProcesoParte.Rol.PASIVA)

            if limpiar_texto(proyecto_o_motivo):
                tipo_proceso, _ = TipoProceso.objects.get_or_create(
                    categoria=categoria, nombre=limpiar_texto(proyecto_o_motivo)[:255]
                )
                proceso.tipo_proceso = tipo_proceso
                proceso.save(update_fields=["tipo_proceso"])

        # Guarda cada estado distinto encontrado en el grupo como entrada de historial
        estados_vistos = set()
        for fila in filas_grupo:
            fila_completa = list(fila) + [None] * (9 - len(fila))
            texto = limpiar_texto(fila_completa[7])
            if texto and texto not in estados_vistos:
                estados_vistos.add(texto)
                HistorialEstado.objects.create(proceso=proceso, estado_nuevo=estado_defecto, observacion=texto)

        return proceso

    def _vincular_partes(self, proceso, texto, rol):
        # Litisconsorcio: "1. Fulano - 2. Mengano - 3. Zutano"
        nombres = re.split(r"\s*-\s*(?=\d+\.)", texto)
        for nombre in nombres:
            nombre = re.sub(r"^\d+\.\s*", "", nombre).strip(" .")
            if not nombre or len(nombre) < 3:
                continue
            parte, _ = Parte.objects.get_or_create(nombre=nombre[:500])
            ProcesoParte.objects.create(proceso=proceso, parte=parte, rol=rol)